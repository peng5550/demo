from openpyxl import load_workbook
from tkinter.messagebox import *
from tkinter import scrolledtext, filedialog
from mttkinter import mtTkinter as mtk
import threading
from datetime import datetime
import settings
import locale
import copy
from openpyxl.utils import get_column_letter

locale.setlocale(locale.LC_CTYPE, 'chinese')


class Cnumber:
    cdict = {}
    gdict = {}
    xdict = {}

    def __init__(self):
        self.cdict = {1: u'', 2: u'拾', 3: u'佰', 4: u'仟'}
        self.xdict = {1: u'元', 2: u'万', 3: u'亿', 4: u'兆'}  # 数字标识符
        self.gdict = {0: u'零', 1: u'壹', 2: u'贰', 3: u'叁', 4: u'肆', 5: u'伍', 6: u'陆', 7: u'柒', 8: u'捌', 9: u'玖'}

    @staticmethod
    def csplit(cdata):  # 拆分函数，将整数字符串拆分成[亿，万，仟]的list
        g = len(cdata) % 4
        csdata = []
        lx = len(cdata) - 1
        if g > 0:
            csdata.append(cdata[0:g])
        k = g
        while k <= lx:
            csdata.append(cdata[k:k + 4])
            k += 4
        return csdata

    def cschange(self, cki):  # 对[亿，万，仟]的list中每个字符串分组进行大写化再合并
        lenki = len(cki)
        lk = lenki
        chk = u''
        for i in range(lenki):
            if int(cki[i]) == 0:
                if i < lenki - 1:
                    if int(cki[i + 1]) != 0:
                        chk = chk + self.gdict[int(cki[i])]
            else:
                chk = chk + self.gdict[int(cki[i])] + self.cdict[lk]
            lk -= 1
        return chk

    def cwchange(self, data):
        cdata = str(data).split('.')

        cki = cdata[0]
        ckj = cdata[1]

        chk = u''
        cski = self.csplit(cki)  # 分解字符数组[亿，万，仟]三组List:['0000','0000','0000']
        ikl = len(cski)  # 获取拆分后的List长度
        # 大写合并
        for i in range(ikl):
            if self.cschange(cski[i]) == '':  # 有可能一个字符串全是0的情况
                chk = chk + self.cschange(cski[i])  # 此时不需要将数字标识符引入
            else:
                chk = chk + self.cschange(cski[i]) + self.xdict[ikl - i]  # 合并：前字符串大写+当前字符串大写+标识符
        # 处理小数部分
        lenkj = len(ckj)
        if lenkj == 1:  # 若小数只有1位
            if int(ckj[0]) == 0:
                chk = chk + u'整'
            else:
                chk = chk + self.gdict[int(ckj[0])] + u'角整'
        else:  # 若小数有两位的四种情况
            if int(ckj[0]) == 0 and int(ckj[1]) != 0:
                chk = chk + u'零' + self.gdict[int(ckj[1])] + u'分'
            elif int(ckj[0]) == 0 and int(ckj[1]) == 0:
                chk = chk + u'整'
            elif int(ckj[0]) != 0 and int(ckj[1]) != 0:
                chk = chk + self.gdict[int(ckj[0])] + u'角' + self.gdict[int(ckj[1])] + u'分'
            else:
                chk = chk + self.gdict[int(ckj[0])] + u'角整'
        return chk


class Application:

    def __init__(self, master):
        self.root = master
        self.root.geometry("500x300")
        self.root.title("Tool")
        self.__createUI()
        self.pt = Cnumber()
        self.addLog("请导入Excel数据.")

    def addLog(self, msg):
        self.logText.insert(mtk.END, "{} {}\n".format(datetime.now().strftime("%H:%M:%S"), msg))
        self.logText.yview_moveto(1.0)

    def __createUI(self):
        # 日志信息
        self.logBox = mtk.LabelFrame(self.root, text="日志信息", fg="blue")
        self.logBox.place(x=20, y=20, width=250, height=260)
        self.logText = scrolledtext.ScrolledText(self.logBox, fg="green")
        self.logText.place(x=20, y=15, width=220, height=210)
        # Excel数据
        self.excelDataBox = mtk.LabelFrame(self.root, text="当前数据", fg="blue")
        self.excelDataBox.place(x=290, y=20, width=180, height=80)
        self.dataNum = mtk.Label(self.excelDataBox, text="当前数据数量：")
        self.dataNum.place(x=10, y=10, width=100, height=30)
        self.dataNumText = mtk.Label(self.excelDataBox, text=0)
        self.dataNumText.place(x=115, y=10, width=30, height=30)

        # taskInfo
        self.taskInfoBox = mtk.LabelFrame(self.root, text="启动任务", fg="blue")
        self.taskInfoBox.place(x=290, y=120, width=180, height=160)
        self.excelExportInBtn = mtk.Button(self.taskInfoBox, text="导入Excel",
                                           command=lambda: self.thread_it(self.excelExportIn))
        self.excelExportInBtn.place(x=15, y=20, width=80, height=35)
        self.excelExportOutBtn = mtk.Button(self.taskInfoBox, text="导出Excel",
                                            command=lambda: self.thread_it(self.excelExportOut))
        self.excelExportOutBtn.place(x=15, y=80, width=80, height=35)
        self.taskStartBtn = mtk.Button(self.taskInfoBox, text="开始", command=lambda: self.thread_it(self.start))
        self.taskStartBtn.place(x=110, y=20, width=50, height=90)

    def excelExportIn(self):
        excelPath = filedialog.askopenfilename(title=u"选择文件")
        if excelPath:
            try:
                wb = load_workbook(excelPath)
                ws = wb.active
                self.excelDataIndex = [index for index, data in enumerate(list(ws.values)[2:]) if data[0]]
                self.excelData = [list(ws.values)[2:][i:self.excelDataIndex[index + 1]] for index, i in
                                  enumerate(self.excelDataIndex[:-1])]
                self.dataNumText.configure(text=len(self.excelDataIndex))
                self.addLog(f"导入Excel成功, 共发现{len(self.excelDataIndex)}条数据.")
            except Exception as e:
                showerror("错误信息", "请导入正确的Excel!")
        else:
            showerror("错误信息", "请导入Excel!")

    def excelExportOut(self):
        excelPath = filedialog.asksaveasfilename(title=u'保存文件', filetypes=[("xlsx", ".xlsx")]) + ".xlsx"
        if excelPath.strip(".xlsx"):
            self.wb.save(excelPath)
        else:
            showerror("错误信息", "请输入正确的Excel名称!")
            return
        showinfo("提示信息", "导出文件成功!")
        self.addLog("当前任务完成!")

    def copySheet(self, wb, copysheetname, sheetname="Template"):
        sheet = wb[sheetname]
        sheet2 = wb.create_sheet(copysheetname, 0)
        # tab颜色
        sheet2.sheet_properties.tabColor = sheet.sheet_properties.tabColor
        # 开始处理合并单元格形式为“(<CellRange A1：A4>,)，替换掉(<CellRange 和 >,)' 找到合并单元格
        wm = list(sheet.merged_cells)
        if len(wm) > 0:
            for i in range(0, len(wm)):
                cell2 = str(wm[i]).replace('(<CellRange ', '').replace('>,)', '')
                sheet2.merge_cells(cell2)

        for i, row in enumerate(sheet.iter_rows()):
            sheet2.row_dimensions[i + 1].height = sheet.row_dimensions[i + 1].height
            for j, cell in enumerate(row):
                sheet2.column_dimensions[get_column_letter(j + 1)].width = sheet.column_dimensions[
                    get_column_letter(j + 1)].width
                sheet2.cell(row=i + 1, column=j + 1, value=cell.value)

                # 设置单元格格式
                source_cell = sheet.cell(i + 1, j + 1)
                target_cell = sheet2.cell(i + 1, j + 1)
                target_cell.fill = copy.copy(source_cell.fill)
                if source_cell.has_style:
                    target_cell._style = copy.copy(source_cell._style)
                    target_cell.font = copy.copy(source_cell.font)
                    target_cell.border = copy.copy(source_cell.border)
                    target_cell.fill = copy.copy(source_cell.fill)
                    target_cell.number_format = copy.copy(source_cell.number_format)
                    target_cell.protection = copy.copy(source_cell.protection)
                    target_cell.alignment = copy.copy(source_cell.alignment)

        return wb

    def getExcelContent(self, excelData):
        deliveryData = [(i[13].month, i[13].day, i[14], i[15], i[16], i[15] * float(i[16])) for i in excelData if
                        i[13]]
        paymentsData = [(i[19].month, i[19].day, i[20], i[21]) for i in excelData if i[19]]
        data = {
            "gysName": excelData[0][0],
            "gysSalesman": excelData[0][1],
            "gysPhone": excelData[0][2],
            "gysFax": excelData[0][3],
            "gysAddress": excelData[0][4],
            "gysBankNo": excelData[0][0],
            "khName": excelData[0][0],
            "khContacts": excelData[0][7],
            "khPhone": excelData[0][8],
            "khFax": excelData[0][9],
            "khAddress": excelData[0][10],
            "date": f"对账期间：{datetime.strftime(excelData[0][11], '%Y年%m月%d日')}-{datetime.strftime(excelData[0][12], '%Y年%m月%d日')}",
            "deliveryData": deliveryData,
            "amountReceivable": sum([i[-1] for i in deliveryData]),
            "paymentsData": paymentsData,
            "paymentsTotal": sum([i[-1] for i in paymentsData]),
            "lastArrears": excelData[0][23]
        }

        return data

    def loadTemplate(self):
        wb = load_workbook("./template.xlsx")
        ws = wb.get_sheet_by_name("Template")
        return wb, ws

    def changeData(self, ws, data):
        ws[settings.billTitle] = f"{data['gysName']}对账单"
        ws[settings.billDate] = data["date"]
        ws[settings.gysName] = f"供应商：{data['gysName']}"
        ws[settings.gysSalesman] = f"联系人：{data['gysSalesman']}"
        ws[settings.gysPhone] = f"电话：{data['gysPhone']}"
        ws[settings.gysFax] = f"传真：{data['gysFax']}"
        ws[settings.gysAddress] = f"地址：{data['gysAddress']}"
        ws[settings.gysBankNo] = f"收款账号：{data['gysBankNo']}"
        ws[settings.khName] = f"客户：{data['khName']}"
        ws[settings.khContacts] = f"联系人：{data['khContacts']}"
        ws[settings.khPhone] = f"电话：{data['khPhone']}"
        ws[settings.khFax] = f"传真：{data['khFax']}"
        ws[settings.khAddress] = f"地址：{data['khAddress']}"
        ws[settings.amountReceivable] = data["amountReceivable"]
        ws[settings.paymentsTotal] = data["paymentsTotal"]
        ws[settings.lastArrears] = data["lastArrears"]
        totalsArrears = float(data["amountReceivable"] + data["lastArrears"] - data["paymentsTotal"])
        ws[settings.totalsArrears] = totalsArrears
        arrearsText = self.pt.cwchange('%.2f' % totalsArrears)
        arrearsFormat = format(totalsArrears, ",")
        if len(arrearsFormat.split(".")[-1]) < 2:
            arrearsFormat += "0"
        ws[
            settings.text] = f"截止至{data['date'].split('-')[-1]}，贵公司共欠我公司货款：{arrearsText}（￥{arrearsFormat}）。请核对后在三天内盖章回传，否则视为无误。感谢贵公司一直以来的支持和惠顾！"
        ws[settings.gysName_] = f"{data['gysName']}（签字盖章）"
        ws[settings.khName_] = f"{data['khName']}（签字盖章）"
        ws[settings.billDate_1] = data['date'].split('-')[-1]
        ws[settings.billDate_2] = data['date'].split('-')[-1]
        # 插入数据
        deliveryData = data["deliveryData"]
        iindex = ["b", "c", "d", "f", "g", "h"]
        for index, line in enumerate(deliveryData):
            for i in range(6):
                ws[f"{iindex[i]}{settings.data_index1 + index}"] = line[i]
        paymentsData = data["paymentsData"]
        iindex = ["b", "c", "d", "h"]
        for index, line in enumerate(paymentsData):
            for i in range(4):
                ws[f"{iindex[i]}{settings.data_index2 + index}"] = line[i]

        return ws

    @staticmethod
    def thread_it(func, *args):
        t = threading.Thread(target=func, args=args)
        t.setDaemon(True)
        t.start()

    def start(self):
        if not self.excelData:
            showerror("错误信息", "当前无数据, 请导入数据.")
            return
        try:
            self.wb, ws = self.loadTemplate()
            for index, data in enumerate(self.excelData):
                self.wb = self.copySheet(self.wb, f"ok_{index}")
                data00 = self.getExcelContent(data)
                ws = self.wb[f"ok_{index}"]
                self.changeData(ws, data00)
            self.addLog("处理完成, 请导出Excel.")
        except Exception as e:
            self.addLog(e.args)

if __name__ == '__main__':
    root = mtk.Tk()
    Application(root)
    root.mainloop()
